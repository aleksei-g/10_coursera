import requests
import random
from xml.etree import ElementTree
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border,  Side, PatternFill
import os.path
import argparse
import sys


URL_TO_COURSERA_XML = 'https://www.coursera.org/sitemap~www~courses.xml'
NUMBER_ANALYZED_COURSES = 20
TEXT_FOR_LINK = 'подробнее'
COLUMNS_ORDER = {
                 'name_course': [1, 'Имя'],
                 'language_course': [2, 'Язык'],
                 'starts_course': [3, 'Дата начала'],
                 'number_weeks_course': [4, 'Продолжительность (недель)'],
                 'rating_course': [5, 'Рейтинг'],
                 'course_url': [6, 'URL']
                }


def create_parser():
    parser = argparse.ArgumentParser(description='Скрипт выполняет сбор \
                                     информации о разных курсах на Курсере.')
    parser.add_argument('-f', '--file', metavar='ФАЙЛ',
                        help='Имя файла для выгрузки информации.')
    return parser


def check_type_file(filepath):
    root, ext = os.path.splitext(filepath)
    return ext in ['.xls', '.xlsx']


def check_filepath(filepath):
    dir_name, file_name = os.path.split(os.path.abspath(filepath))
    return os.path.exists(dir_name)


def get_courses_list():
    response = requests.get(URL_TO_COURSERA_XML).content
    tree = ElementTree.fromstring(response)
    courses_list = []
    for num in range(NUMBER_ANALYZED_COURSES):
        courses_list.append(random.choice(tree)[0].text)
    return courses_list


def get_tag_text(tag):
    return tag.text if tag else None


def get_name_course(soup):
    return get_tag_text(soup.find('div', {'class': 'title display-3-text'}))


def get_rating_course(soup):
    return get_tag_text(soup.find('div', {'class':
                                          'ratings-text bt3-visible-xs'}))


def get_number_weeks_course(soup):
    return len(soup.findAll('div', {'class': 'week'}))


def get_starts_course(soup):
    text_json = get_tag_text(soup.find('script', {'type':
                                       'application/ld+json'}))
    if text_json:
        json_data = json.loads(text_json)
        return json_data['hasCourseInstance'][0]['startDate']


def get_language_course(soup):
    table = soup.find('table', {'class':
                                'basic-info-table bt3-table bt3-table-striped '
                                'bt3-table-bordered bt3-table-responsive'})
    if table:
        all_cols = []
        for row in table:
            for col in row.find_all('td'):
                all_cols.append(col.text)
        col_name_language = all_cols.index('Language')
        if col_name_language is not None:
            return all_cols[col_name_language+1]


def get_site_page(url, payload=None):
    response = requests.get(url, payload)
    response.encoding = 'utf-8'
    return response.text


def get_course_info(course_url):
    page = get_site_page(course_url)
    soup = BeautifulSoup(page, 'lxml')
    course_info = {}
    course_info['name_course'] = get_name_course(soup)
    course_info['rating_course'] = get_rating_course(soup)
    course_info['number_weeks_course'] = get_number_weeks_course(soup)
    course_info['starts_course'] = get_starts_course(soup)
    course_info['language_course'] = get_language_course(soup)
    course_info['course_url'] = course_url
    return course_info


def get_thin_border():
    return Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  top=Side(style='thin'),
                  bottom=Side(style='thin'))


def get_darkgray_fill():
    return PatternFill(start_color='A9A9A9',
                       end_color='A9A9A9',
                       fill_type='solid')


def get_lightgray_fill():
    return PatternFill(start_color='D3D3D3',
                       end_color='D3D3D3',
                       fill_type='solid')


def output_courses_info_to_xlsx(courses_info, filepath):
    wb = Workbook()
    sheet = wb.active
    for item in COLUMNS_ORDER.items():
        cell = sheet.cell(row=1, column=item[1][0])
        cell.value = item[1][1]
        cell.font = Font(bold=True)
        cell.border = get_thin_border()
        cell.fill = get_darkgray_fill()
    for row, course in enumerate(courses_info, start=2):
        for item in course.items():
            cell = sheet.cell(row=row, column=COLUMNS_ORDER[item[0]][0])
            if item[0] == 'course_url':
                cell.value = '=HYPERLINK("%s","%s")' % (item[1], TEXT_FOR_LINK)
            else:
                cell.value = item[1]
            cell.border = get_thin_border()
            if row % 2:
                cell.fill = get_lightgray_fill()
    sheet.column_dimensions['A'].width = 50
    wb.save(filepath)
    return True


if __name__ == '__main__':
    parser = create_parser()
    namespace = parser.parse_args()
    if namespace.file:
        filepath = namespace.file
    else:
        filepath = input('Введите имя файла для выгрузки данных '
                         'в формате "xls" или "xlsx":\n')
    if not check_filepath(filepath):
        print('Каталог не существует!')
        sys.exit(1)
    else:
        if not check_type_file(filepath):
            print('Файл должен быть формата "xls" или "xlsx".')
            sys.exit(1)
    courses_list = get_courses_list()
    courses_info = [get_course_info(course) for course in courses_list]
    if output_courses_info_to_xlsx(courses_info, filepath):
        print('Информация о курсах выгружена в файл "%s"' %
              os.path.abspath(filepath))
